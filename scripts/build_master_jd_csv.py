#!/usr/bin/env python3
"""Build scripts/positions.master.csv from the client's Example Master JD.xlsx.

The Master JD sheet has columns:
  A ชื่อตำแหน่ง  ("English (ไทย)")   B ระดับตำแหน่ง  (Staff/Supervisor/Manager/Officer)
  C หน้าที่รับผิดชอบ (responsibilities)  D คุณสมบัติ (qualifications)  E Highlighted รายละเอียดงาน

It emits the CSV consumed by `backend/cmd/importref` (importPositions), with the
extra JD-text columns the scorer now reads. Derivations are deterministic and
printed as a summary table so a human can eyeball/edit before importing.

Education ordinals match backend/internal/scoring/rules.go:
  1 = ม.6/ปวช/high-school   2 = ปวส/diploma   3 = bachelor   4 = master

Usage: python3 scripts/build_master_jd_csv.py [path/to/Example Master JD.xlsx]
Output: scripts/positions.master.csv
"""
import csv
import os
import re
import sys

import openpyxl

DEFAULT_XLSX = (
    "/Users/nex/My Drive/Nexto Company/All of Works/Nexto Software/2026/ERT/CP/"
    "requirementcpaxtranexto/Example Master JD.xlsx"
)
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "positions.master.csv")

# Education keyword → ordinal. Order matters (check most-specific first).
EDU_RULES = [
    (re.compile(r"ปริญญาโท|ป\.โท|master"), 4),
    (re.compile(r"ปริญญาตรี|ป\.ตรี|ปริญญา|bachelor|ตรี"), 3),
    (re.compile(r"ปวส|อนุปริญญา|diploma|ปวส\."), 2),
    (re.compile(r"ปวช|ม\.6|ม\.ปลาย|มัธยมศึกษาตอนปลาย|มัธยม|high school|secondary"), 1),
]
# Level → fallback when the qualifications text names no education / experience.
EDU_BY_LEVEL = {"manager": 3, "supervisor": 2, "officer": 1, "staff": 1}
EXP_BY_LEVEL = {"manager": 24, "supervisor": 12, "officer": 0, "staff": 0}

# "ประสบการณ์ ... 3-4 ปี" → take the lower bound (3). Tolerates a gap before the years.
EXP_NUM = re.compile(r"ประสบการณ์[^\n•]{0,60}?(\d+)\s*[-–]?\s*\d*\s*ปี")
# Experience phrased as a plus, not a hard requirement.
EXP_OPTIONAL = re.compile(r"พิจารณาเป็นพิเศษ|จะพิจารณา|พร้อมเรียนรู้|หากมีประสบการณ์")

LEVEL_WORDS = {"manager", "supervisor", "staff", "officer", "admin", "executive"}


def split_title(raw: str):
    """'English (ไทย)' → (title_en, title_th). EN-only rows fall back th=en."""
    raw = " ".join(str(raw or "").split())
    m = re.match(r"^(.*?)\((.*)\)\s*$", raw)
    if m:
        en = m.group(1).strip(" -–")
        th = m.group(2).strip()
        if not en:  # parenthetical-only
            en = th
        return en, th
    return raw, raw


def ps_code(title_en: str, used: set) -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "_", title_en).strip("_").upper() or "POSITION"
    code = base
    i = 2
    while code in used:
        code = f"{base}_{i}"
        i += 1
    used.add(code)
    return code


def min_education(qual: str, level: str) -> int:
    for pat, ord_ in EDU_RULES:
        if pat.search(qual):
            return ord_
    return EDU_BY_LEVEL.get(level, 1)


def min_experience(qual: str, level: str) -> int:
    m = EXP_NUM.search(qual)
    if m and not EXP_OPTIONAL.search(qual):
        return int(m.group(1)) * 12
    if EXP_OPTIONAL.search(qual) and level in ("staff", "officer"):
        return 0
    return EXP_BY_LEVEL.get(level, 0)


def keywords(title_en: str) -> str:
    toks = re.split(r"[^A-Za-z0-9]+", title_en.lower())
    kw = [t for t in toks if t and t not in LEVEL_WORDS and not t.isdigit() and len(t) > 1]
    # de-dup, keep order
    seen, out = set(), []
    for t in kw:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return "|".join(out)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    ws = openpyxl.load_workbook(path, data_only=True).active

    used_codes: set = set()
    rows = []
    for r in range(2, ws.max_row + 1):
        title_raw = ws.cell(r, 1).value
        if not title_raw or not str(title_raw).strip():
            continue
        level_raw = str(ws.cell(r, 2).value or "").strip()
        level = level_raw.lower()
        resp = str(ws.cell(r, 3).value or "").strip()
        qual = str(ws.cell(r, 4).value or "").strip()
        highlight = str(ws.cell(r, 5).value or "").strip()

        title_en, title_th = split_title(title_raw)
        # Fold the Highlighted details into responsibilities for a richer JD body.
        responsibilities = resp if not highlight else f"{resp}\n\n{highlight}".strip()

        rows.append({
            "title_th": title_th,
            "title_en": title_en,
            "level": level_raw or "Staff",
            "ps_position_code": ps_code(title_en, used_codes),
            "min_education_level": min_education(qual, level),
            "min_experience_months": min_experience(qual, level),
            "keywords": keywords(title_en),
            "responsibilities": responsibilities,
            "qualifications": qual,
        })

    cols = ["title_th", "title_en", "level", "ps_position_code",
            "min_education_level", "min_experience_months", "keywords",
            "responsibilities", "qualifications"]
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        w.writerows(rows)

    print(f"Wrote {len(rows)} positions → {OUT}\n")
    print(f"{'#':>2}  {'CODE':<26} {'LVL':<10} {'EDU':>3} {'EXPmo':>5}  TITLE")
    for i, row in enumerate(rows, 1):
        print(f"{i:>2}  {row['ps_position_code']:<26} {row['level']:<10} "
              f"{row['min_education_level']:>3} {row['min_experience_months']:>5}  "
              f"{row['title_en'][:40]}")


if __name__ == "__main__":
    main()
